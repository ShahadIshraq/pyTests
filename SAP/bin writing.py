import openpyxl as op

# reading the excel file
import os

fileName = 'State.xlsx'
wb = op.load_workbook(fileName)
sheet = wb.get_sheet_by_name("Sheet1")

print("Number of columns to parse: "+str(sheet.max_column-5-5))
print("Number of rows to parse: "+str(sheet.max_row-2))

for r in range(3, sheet.max_row+1):
    print("****************")
    print("In row : " + str(r))
    print("****************")
    temp = 0
    pos = 7
    for c in range(6, 14):
        content = sheet.cell(row=r, column=c).value
        if content is not None:
            value = int(content)
            if value == 1:
                temp = temp | 1 << pos
        pos = pos - 1
    sheet.cell(row=r, column=c+1).value = format(temp, '02X')
    print("Bin1 = "+sheet.cell(row=r, column=c+1).value)
    temp = 0
    pos = 7
    for c in range(15, 23):
        content = sheet.cell(row=r, column=c).value
        if content is not None:
            value = int(content)
            if value == 1:
                temp = temp | 1 << pos
        pos = pos - 1
    sheet.cell(row=r, column=c+1).value = format(temp, '02X')
    print("Bin2 = "+sheet.cell(row=r, column=c+1).value)

    temp = 0
    pos = 7
    for c in range(24, 32):
        content = sheet.cell(row=r, column=c).value
        if content is not None:
            value = int(content)
            if value == 1:
                temp = temp | 1 << pos
        pos = pos - 1
    sheet.cell(row=r, column=c+1).value = format(temp, '02X')
    print("Bin3 = "+sheet.cell(row=r, column=c+1).value)

    temp = 0
    pos = 7
    for c in range(33, 41):
        content = sheet.cell(row=r, column=c).value
        if content is not None:
            value = int(content)
            if value == 1:
                temp = temp | 1 << pos
        pos = pos - 1
    sheet.cell(row=r, column=c+1).value = format(temp, '02X')
    print("Bin4 = "+sheet.cell(row=r, column=c+1).value)

    temp = 0
    pos = 7
    for c in range(42, 50):
        content = sheet.cell(row=r, column=c).value
        if content is not None:
            value = int(content)
            if value == 1:
                temp = temp | 1 << pos
        pos = pos - 1
    sheet.cell(row=r, column=c+1).value = format(temp, '02X')
    print("Bin5 = "+sheet.cell(row=r, column=c+1).value)
    print()

wb.save("changedState.xlsx")
wb = op.load_workbook("changedState.xlsx")
sheet = wb.get_sheet_by_name("Sheet1")


directory = "./bins/"
if not os.path.exists(directory):
    os.makedirs(directory)

c = 14
file = open(directory+"bin1.bin", 'w')
y = []
for r in range(3, sheet.max_row+1):
    content = sheet.cell(row=r, column=c).value + "\n"
    y.append(content)
file.writelines(y)
file.close()

c = 23
file = open(directory+"bin2.bin", 'w')
y = []
for r in range(3, sheet.max_row+1):
    content = sheet.cell(row=r, column=c).value + "\n"
    y.append(content)
file.writelines(y)
file.close()

c = 32
file = open(directory+"bin3.bin", 'w')
y = []
for r in range(3, sheet.max_row+1):
    content = sheet.cell(row=r, column=c).value + "\n"
    y.append(content)
file.writelines(y)
file.close()


c = 41
file = open(directory+"bin4.bin", 'w')
y = []
for r in range(3, sheet.max_row+1):
    content = sheet.cell(row=r, column=c).value + "\n"
    y.append(content)
file.writelines(y)
file.close()

c = 50
file = open(directory+"bin5.bin", 'w')
y = []
for r in range(3, sheet.max_row+1):
    content = sheet.cell(row=r, column=c).value + "\n"
    y.append(content)
file.writelines(y)
file.close()