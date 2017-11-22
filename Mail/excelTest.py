import openpyxl as op

wb = op.load_workbook('excel.xlsx')

sheet = wb.get_active_sheet();

for r in range (2,57) :
    cell = sheet.cell(row=r , column=3).value
    if (cell == 'Buet' or cell == 'BUET' or cell == 'buet' or cell == 'Bangladesh University of Engineering and Technology [BUET]' or cell == 'Bangladesh University of Engineering and Technology'):
        print(sheet.cell(row=r , column=2).value + '\n')

