import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text  import MIMEText
import openpyxl as op
from string import Template

count = 0

# creating the mail body
f = open("mail.txt", "r")
mailBody = f.read()
f.close()
s = Template(mailBody)
# print(s.substitute(name='Shahad Ishraq', cName='Intro to arduino'))

# reading the excel file
wb = op.load_workbook('excel.xlsx')
sheet = wb.get_active_sheet();

# setting up the server
fromaddr = "teaminterplanetar@gmail.com"
server = smtplib.SMTP('smtp.gmail.com', 587)
server.starttls()
server.login(fromaddr, "interpl@netar")
print("Logged in")

for r in range (56, 57):
    # cell = sheet.cell(row=r , column=3).value
    name = sheet.cell(row=r, column=2).value
    name = name.title()
    toaddr = sheet.cell(row=r, column=4).value
    courses = sheet.cell(row=r, column=6).value
    courses = courses.split(', ')
    if 'CAD' in courses:
        courses.remove('CAD')
    if 'SolidWorks' in courses:
        courses.remove('SolidWorks')
    if len(courses) == 0:
        continue
    # creating the message
    msg = MIMEMultipart()
    msg['From'] = fromaddr
    msg['To'] = toaddr
    msg['Subject'] = "Orientation class for the course " + courses[0]
    if courses[0] == 'Introduction to programming with C' :
        body = s.substitute(name=name, cName=courses[0], time='10')
    else:
        body = s.substitute(name=name, cName=courses[0], time='11')

    count += 1
    msg.attach(MIMEText(body, 'plain'))
    text = msg.as_string()
    server.sendmail(fromaddr, toaddr, text)
    print('To: '+name+' Address: '+toaddr+' Subject: '+msg['Subject'])

    if len(courses) == 2 :
        msg = MIMEMultipart()
        msg['From'] = fromaddr
        msg['To'] = toaddr
        msg['Subject'] = "Orientation class for the course " + courses[1]
        if courses[1] == 'Introduction to programming with C':
            body = s.substitute(name=name, cName=courses[1], time='10')
        else:
            body = s.substitute(name=name, cName=courses[1], time='11')
        msg.attach(MIMEText(body, 'plain'))
        text = msg.as_string()
        server.sendmail(fromaddr, toaddr, text)
        print('To: ' + name + ' Address: ' + toaddr + ' Subject: ' + msg['Subject'])
        count += 1


server.quit()
print(count)
